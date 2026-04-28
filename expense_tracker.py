from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
import csv
import json
import os
from colorama import init, Fore, Style
init(autoreset=True)
EXPENSES_FILE ="expenses.csv"
BUDGET_FILE = "budget.json"

def initialize_csv():
    if not os.path.exists(EXPENSES_FILE):
        with open(EXPENSES_FILE, mode='w',newline='') as file:
            writer = csv.writer(file)
            writer.writerow(["date","category","description","amount"])

def add_expense():
    print("\n--- Add New Expense ---")
    date = input("Enter date (DD/MM/YYYY): ")
    
    categories = ["Food", "Transport", "Shopping", "Groceries", "Other"]
    print("\nCategories:")
    for i, cat in enumerate(categories, 1):
        print(f"{i}. {cat}")
    cat_choice = input("Choose category (1-5): ")
    
    try:
        category = categories[int(cat_choice) - 1]
    except:
        print(Fore.RED + "Invalid choice, setting to Other")
        category = "Other"
    
    description = input("Enter description: ")
    amount = input("Enter amount (€): ")

    with open(EXPENSES_FILE, mode='a', newline='') as file:
        writer = csv.writer(file)
        writer.writerow([date, category, description, amount])

    print(Fore.GREEN + "\n Expense added successfully!")

def view_expenses():
    print("\n--- All Expenses ---")
    with open(EXPENSES_FILE, mode='r') as file:
        reader = csv.DictReader(file)
        expenses = list(reader)
    if not expenses:
        print(Fore.RED + "No expenses found!")
        return
    print(f"\n{'No.':<5}{'Date':<15}{'Category':<15}{'Description':<25}{'Amount(€)':<10}")
    print("-"*70)
    for i, row in enumerate(expenses, 1):
        print(f"{i:<5} {row['date']:<15} {row['category']:<15} {row['description']:<25} {row['amount']:<10}")
def load_budget():
    if os.path.exists(BUDGET_FILE):
        with open(BUDGET_FILE,"r") as file:
            return json.load(file)
    return {"total": 0, "categories":{}}

def set_budget():
    print("\n---Set your Budget ---")
    total= float(input("Enter your total monthly budget(€): "))
    categories = ["Food", "Transport", "Shopping", "Groceries","Other"]
    category_budgets = {}
    print("\nNow set a limit for each category:")
    for category in categories:
        amount = float(input(f"  {category} budget (€): "))
        category_budgets[category] = amount
    category_total = sum(category_budgets.values())

    if category_total > total:
        print(Fore.RED + f"\n Warning: Your category budgets (€{category_total}) exceed your total budget (€{total})!")
        confirm = input("Do you still want to save? (yes/no): ")
        if confirm.lower() != "yes":
            print(Fore.RED + "Budget not saved. Please try again.")
            return
    
    
    budget = {
        "total": total,
        "categories": category_budgets
    }
    
    with open(BUDGET_FILE, "w") as file:
        json.dump(budget, file, indent=4)
    
    print(Fore.GREEN + "\n Budget saved successfully!")

def show_summary():
    print("\n--- Spending Summary ---")
    
    # Load expenses
    with open(EXPENSES_FILE, mode='r') as file:
        reader = csv.DictReader(file)
        expenses = list(reader)
    
    if not expenses:
        print(Fore.RED + "No expenses found!")
        return
    
    # Load budget
    budget = load_budget()
    total_budget = budget["total"]
    category_budgets = budget["categories"]
    
    # Calculate total spent
    total_spent = sum(float(row["amount"]) for row in expenses)
    
    # Calculate spent per category
    category_spent = {}
    for row in expenses:
        cat = row["category"]
        category_spent[cat] = category_spent.get(cat, 0) + float(row["amount"])
    
    # Total budget summary
    print(f"\n{'='*40}")
    print(f"Total Budget:  €{total_budget:.2f}")
    print(f"Total Spent:   €{total_spent:.2f}")
    print(f"Remaining:     €{total_budget - total_spent:.2f}")
    
    percentage = (total_spent / total_budget * 100) if total_budget > 0 else 0
    
    if percentage >= 100:
        print(Fore.RED + f"  You have exceeded your total budget! ({percentage:.1f}%)")
    elif percentage >= 75:
        print(Fore.YELLOW + f"  Warning: You have used {percentage:.1f}% of your total budget!")
    else:
        print(Fore.GREEN + f" You are within budget! ({percentage:.1f}% used)")
    
    # Category breakdown
    print(f"\n{'Category':<15}{'Spent':>10}{'Budget':>10}{'Remaining':>12}{'Status'}")
    print("-" * 60)
    
    for category in ["Food", "Transport", "Shopping", "Groceries", "Other"]:
        spent = category_spent.get(category, 0)
        cat_budget = category_budgets.get(category, 0)
        remaining = cat_budget - spent
        percentage_cat = (spent / cat_budget * 100) if cat_budget > 0 else 0
        
        if percentage_cat >= 100:
            status = Fore.RED + "EXCEEDED"
        elif percentage_cat >= 75:
            status = Fore.YELLOW + "WARNING"
        else:
            status = Fore.GREEN + "OK"
        
        print(f"{category:<15}€{spent:>8.2f}  €{cat_budget:>7.2f}  €{remaining:>9.2f}  {status}")
    
    # Most expensive category
    if category_spent:
        most_expensive = max(category_spent, key=category_spent.get)
        print(f"\n Most spent category: {most_expensive} (€{category_spent[most_expensive]:.2f})")
def generate_excel_report():
    print("\n--- Generating Excel Report ---")
    
    # Load expenses
    with open(EXPENSES_FILE, mode='r') as file:
        reader = csv.DictReader(file)
        expenses = list(reader)
    
    if not expenses:
        print(Fore.RED + "No expenses found!")
        return
    
    # Load budget
    budget = load_budget()
    category_budgets = budget["categories"]
    
    # Create workbook
    wb = Workbook()
    
    # ── Sheet 1: All Expenses ──
    ws1 = wb.active
    ws1.title = "All Expenses"
    
    # Header style
    header_fill = PatternFill(start_color="2F75B6", end_color="2F75B6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = ["No.", "Date", "Category", "Description", "Amount (€)"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add expenses
    for i, row in enumerate(expenses, 2):
        ws1.cell(row=i, column=1, value=i-1)
        ws1.cell(row=i, column=2, value=row["date"])
        ws1.cell(row=i, column=3, value=row["category"])
        ws1.cell(row=i, column=4, value=row["description"])
        ws1.cell(row=i, column=5, value=float(row["amount"]))
    
    # Column widths
    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 15
    ws1.column_dimensions["C"].width = 15
    ws1.column_dimensions["D"].width = 30
    ws1.column_dimensions["E"].width = 12
    
    # ── Sheet 2: Category Summary ──
    ws2 = wb.create_sheet("Budget Summary")
    
    # Headers
    headers2 = ["Category", "Spent (€)", "Budget (€)", "Remaining (€)", "% Used", "Status"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Calculate category spending
    category_spent = {}
    for row in expenses:
        cat = row["category"]
        category_spent[cat] = category_spent.get(cat, 0) + float(row["amount"])
    
    # Color fills
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    amber_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    categories = ["Food", "Transport", "Shopping", "Groceries", "Other"]
    for i, category in enumerate(categories, 2):
        spent = category_spent.get(category, 0)
        cat_budget = category_budgets.get(category, 0)
        remaining = cat_budget - spent
        percentage = (spent / cat_budget * 100) if cat_budget > 0 else 0
        
        if percentage >= 100:
            status = "EXCEEDED"
            fill = red_fill
        elif percentage >= 75:
            status = "WARNING"
            fill = amber_fill
        else:
            status = "OK"
            fill = green_fill
        
        ws2.cell(row=i, column=1, value=category)
        ws2.cell(row=i, column=2, value=round(spent, 2))
        ws2.cell(row=i, column=3, value=cat_budget)
        ws2.cell(row=i, column=4, value=round(remaining, 2))
        ws2.cell(row=i, column=5, value=round(percentage, 1))
        status_cell = ws2.cell(row=i, column=6, value=status)
        
        for col in range(1, 7):
            ws2.cell(row=i, column=col).fill = fill
    
    # Column widths
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws2.column_dimensions[col].width = 15
    
    # ── Bar Chart: Spent vs Budget per Category ──
    bar_chart = BarChart()
    bar_chart.title = "Spent vs Budget by Category"
    bar_chart.y_axis.title = "Amount (€)"
    bar_chart.x_axis.title = "Category"
    bar_chart.style = 10
    bar_chart.width = 20
    bar_chart.height = 12
    
    spent_data = Reference(ws2, min_col=2, min_row=1, max_row=6)
    budget_data = Reference(ws2, min_col=3, min_row=1, max_row=6)
    categories_ref = Reference(ws2, min_col=1, min_row=2, max_row=6)
    
    bar_chart.add_data(spent_data, titles_from_data=True)
    bar_chart.add_data(budget_data, titles_from_data=True)
    bar_chart.set_categories(categories_ref)
    ws2.add_chart(bar_chart, "H1")
    
    # Save
    report_file = "expense_report.xlsx"
    wb.save(report_file)
    print(Fore.GREEN + f"\n✅ Excel report generated: {report_file}")
    print(Fore.GREEN + f"📊 Open it in Excel to see your charts!")
def main() :
    initialize_csv()
    while True:
        print("\n==== Expense Tracker ====")
        print("1. Add expense")
        print("2. View All expenses")
        print("3. Set Budget")
        print("4. Show Summary")
        print("5. Generate Excel Report")
        print("6. Exit!")
        choice = input("\nChoose an option: ")

        if choice == "1":
            add_expense()
        elif choice == "2":
            view_expenses()
        elif choice == "3":
           set_budget()
        elif choice == "4":
            show_summary()
        elif choice == "5":
            generate_excel_report()
        elif choice == "6":
              print("Goodbye!!")
              break
        else:
            print(Fore.RED + "Invalid option, try again.")

if __name__ == "__main__":
    main()